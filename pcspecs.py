import logging
import os
import platform
import subprocess
import sys
#Ensure stdout encoding is UTF-8
if sys.stdout.encoding != 'utf-8':
    sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)
#Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("system_info.log", encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)


def install_and_import(package, module_name=None):
    """Install and import a package."""
    try:
        if not module_name:
            module_name = package
        globals()[module_name] = __import__(module_name)
    except ImportError:
        logging.info(f"Package {package} not found. Installing...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])
        globals()[module_name] = __import__(module_name)
    finally:
        logging.info(f"Package {package} is ready to use.")


install_and_import('setuptools')
install_and_import('psutil')
install_and_import('py-cpuinfo', 'cpuinfo')
install_and_import('openpyxl')
install_and_import('GPUtil')
install_and_import('pywin32', 'win32com')
install_and_import('wmi')

import psutil
import cpuinfo
import GPUtil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import wmi


def get_size(bytes, suffix="B"):
    """Convert bytes to a human-readable format."""
    factor = 1024
    for unit in ["", "K", "M", "G", "T", "P"]:
        if bytes < factor:
            return f"{bytes:.2f}{unit}{suffix}"
        bytes /= factor


def get_system_info():
    """Retrieve system information."""
    os_info = platform.system() + " " + platform.release()
    computer_name = platform.node()
    cpu_info = cpuinfo.get_cpu_info()
    cpu_model = cpu_info['brand_raw']
    cpu_cores = psutil.cpu_count(logical=False)
    cpu_threads = psutil.cpu_count(logical=True)
    cpu_freq = psutil.cpu_freq().current

    virtual_memory = psutil.virtual_memory()
    total_memory_gb = virtual_memory.total / (1024 ** 3)

    total_disk_usage = psutil.disk_usage('/').total / (1024 ** 3)

    gpus = GPUtil.getGPUs()
    gpu_name = gpus[0].name if gpus else 'Built-in'
    gpu_memory_total = gpus[0].memoryTotal if gpus else 'N/A'

    logging.info("System information retrieved successfully.")
    return {
        'os_info': os_info,
        'computer_name': computer_name,
        'cpu_model': cpu_model,
        'cpu_cores': cpu_cores,
        'cpu_threads': cpu_threads,
        'cpu_freq': cpu_freq,
        'total_memory_gb': total_memory_gb,
        'total_disk_usage': total_disk_usage,
        'gpu_name': gpu_name,
        'gpu_memory_total': gpu_memory_total
    }


def get_disk_type():
    """Retrieve disk types using WMI."""
    c = wmi.WMI()
    disk_types = []
    for disk in c.Win32_DiskDrive():
        interface_types = [disk.InterfaceType]
        disk_types.extend(interface_types)

    matching_types = []
    for disk_type in disk_types:
        if disk_type == 'SCSI':
            matching_types.append('SSD')
        elif disk_type == 'USB':
            matching_types.append('USB')
        else:
            matching_types.append('HDD')

    return matching_types


def get_disk_info():
    """Retrieve disk information"""
    disk_info_list = []
    partitions = psutil.disk_partitions()
    disk_types = get_disk_type()

    c = wmi.WMI()
    disk_names = [disk.Model for disk in c.Win32_DiskDrive()]

    for i, partition in enumerate(partitions):
        partition_info = {
            'device': partition.device,
            'disk_type': disk_types[i % len(disk_types)],
            'disk_name': disk_names[i % len(disk_names)]
        }
        try:
            partition_usage = psutil.disk_usage(partition.mountpoint)
            partition_info.update({
                'total_size': get_size(partition_usage.total),
            })
        except PermissionError:
            continue
        disk_info_list.append(partition_info)

    return disk_info_list


def write_to_excel(system_info, disk_info, filename, full_name):
    """Write system and disk information to an Excel file."""
    try:
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'System Info'

        headers = [
            'Full Name', 'OS Info', 'Computer Name', 'CPU Model', 'CPU Cores', 'CPU Threads',
            'CPU Frequency (MHz)', 'Total Memory (GB)', 'Total Disk Usage (GB)',
            'GPU Name', 'GPU Memory Total (MB)', 'Device', 'Disk type', 'Disk Name', 'Total Size'
        ]

        if ws.max_row == 1:
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))

        row = [
            full_name, system_info['os_info'], system_info['computer_name'], system_info['cpu_model'],
            system_info['cpu_cores'],
            system_info['cpu_threads'], system_info['cpu_freq'], system_info['total_memory_gb'],
            system_info['total_disk_usage'], system_info['gpu_name'], system_info['gpu_memory_total']
        ]

        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

        for disk in disk_info:
            disk_row = [''] * 11 + [
                disk['device'], disk['disk_type'], disk['disk_name'], disk.get('total_size', 'N/A')
            ]
            ws.append(disk_row)
            for cell in ws[ws.max_row]:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

        save_path = os.path.join(os.getcwd(), filename)
        wb.save(save_path)
        logging.info(f"System information has been successfully written to '{filename}'")

    except Exception as e:
        logging.error(f"An error occurred while writing to Excel: {e}")


if __name__ == '__main__':
    logging.info("Script started.")
    try:
        full_name = input("Please enter your full name: ")
        system_info = get_system_info()
        disk_info = get_disk_info()
        filename = "system_info.xlsx"
        write_to_excel(system_info, disk_info, filename, full_name)
        logging.info("Script finished successfully.")
    except Exception as e:
        logging.error(f"An error occurred: {e}")
